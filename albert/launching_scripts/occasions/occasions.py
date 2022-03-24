#!/usr/bin/env python3

import datetime

import pandas as pd
import numpy as np
import pytz
from openpyxl import load_workbook

from albert.config.config import smtp_access
from albert.launching_scripts.occasions.occasions_setup import *
from albert.launching_scripts.script_base import BaseScrapinghubScript
from albert.spiders.occasion_main import get_dataframe_for_final_file
from albert.utils.email_utils import send_data
import re


class OccasionsData():
    today = datetime.datetime.now().strftime('%Y-%m-%d')
    fullpath = '/tmp/crawl_seconde_vie_{}'.format(today)

    @staticmethod
    def clean_euro_separates_decimals(price):
        units = price[0].strip().replace('.', '')
        decimals = price[1].strip().replace(',', '.')
        return ','.join([units, decimals])

    def clean_euro_random_price_column(self, price_string):
        final_price_string = None
        if isinstance(price_string, float):
            final_price_string = price_string
        if isinstance(price_string, str):
            price_string = price_string.strip().strip('€').strip().replace(' ', '')
            price_string = self.remove_non_ascii_characters_from_price(price_string)
            price_string = self.remove_alpha_characters_from_price(price_string)
            final_price_string = price_string.replace(' ', '')
            final_price_string = final_price_string.replace(',', '.')
            final_price_string = final_price_string.replace('(', '')
            final_price_string = final_price_string.replace(')', '')
        return final_price_string

    @staticmethod
    def clean_euro_separates_nothing(price):
        return price[0].strip().replace('.', '').replace(',', '.')

    @staticmethod
    def remove_non_ascii_characters_from_price(price):
        return re.sub(r'[^\x00-\x7F]+', '', price)

    @staticmethod
    def remove_alpha_characters_from_price(price):
        return re.sub(r'[A-Za-z]+', '', price)

    @staticmethod
    def remove_non_price_characters(price):
        s = '0123456789,.€'
        return ''.join((filter(lambda x: x in list(s), price)))

    def clean_prices(self, price_string):
        final_price_string = None
        if isinstance(price_string, str):
            separate_euro = self.remove_non_ascii_characters_from_price(price_string)
            separate_euro = self.remove_alpha_characters_from_price(separate_euro)
            separate_euro = self.remove_non_price_characters(separate_euro)
            separate_euro = separate_euro.strip().split('€')
            if len(separate_euro) > 2:
                final_price_string = self.clean_euro_separates_decimals(separate_euro)
            else:
                final_price_string = self.clean_euro_separates_nothing(separate_euro)
            final_price_string = final_price_string.replace(' ', '')
        return final_price_string

    @staticmethod
    def convert_float_to_euros(number):
        return '{}{}'.format(str(number).replace(',', '').replace('.', ','), '€')

    def try_to_guess(self, val):
        if val and isinstance(val, float):
            return str(val)
        return ''

    def decode_job_data(self, job_data):
        try:
            df_list = [
                {key.decode(): val.decode() if val and isinstance(val, bytes) else self.try_to_guess(val) for key, val
                 in d.items()}
                for d in job_data]
        except Exception:
            df_list = None
        return df_list

    @staticmethod
    def init_dataframe(data):
        return pd.DataFrame(data)

    @staticmethod
    def lowercase_dataframe_columns(df):
        return df.columns.str.lower()

    @staticmethod
    def delete_dataframe_columns(df):
        columns_to_delete = ['produit', 'titre', 'product', 'produit apple', 'url', '_type', '_key', 'reference', 'df',
                             'darty_price', 'darty_prix', 'ratio', 'marque', 'delta_darty_amazon', 'type de produit',
                             'darty_codic', 'delta_darty_apple', 'darty_prixttc', 'occasion_second_price_test']
        df.drop(columns=columns_to_delete, inplace=True, errors='ignore')
        return df

    def transform_data(self, job_name, job_data):
        try:
            df_list = self.decode_job_data(job_data)
            df = self.init_dataframe(df_list)
            df.columns = self.lowercase_dataframe_columns(df)
            df = self.delete_dataframe_columns(df)
            availability_dict = {'en stock': 'disponibilité_{}'.format(job_name),
                                 'disponibilité': 'disponibilité_{}'.format(job_name)}
            etat_dict = {'etat': 'etat_{}'.format(job_name), 'état': 'etat_{}'.format(job_name)}
            df.rename(columns=availability_dict, inplace=True)
            df.rename(columns=etat_dict, inplace=True)
            df.drop_duplicates(inplace=True, ignore_index=True)
            df.rename(columns={'price': 'prix'}, inplace=True)
            if 'ean' in df.columns.tolist() and 'darty_ean' in df.columns.tolist():
                df.drop(columns=['darty_ean'], inplace=True, errors='ignore')
            df.rename(columns={'ean': 'EAN'}, inplace=True)
            if 'amazon' in job_name:
                df.drop(columns=['darty_ean'], inplace=True, errors='ignore')
                df.rename(columns={'prix': f'{job_name.upper()}_PrixTTC'}, inplace=True)
            else:
                df.rename(columns={'prix': f'{job_name.upper()}_PrixTTC', 'darty_ean': 'EAN'}, inplace=True)
            df.rename(columns={'prix': f'{job_name.upper()}_PrixTTC'}, inplace=True)
            df[f'{job_name.upper()}_PrixTTC'] = df[f'{job_name.upper()}_PrixTTC'].apply(self.clean_prices)
            return {'df_feuille1': df,
                    'df_feuille_site': df,
                    'site': job_name[:job_name.find('_occ')]}
        except Exception as err:
            print('transform data : {}'.format(err))
            return None

    def add_data_to_excel_file(self, df_dict, writer):
        try:
            writer.book = load_workbook(self.fullpath)
            writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        except Exception as err:
            print('no file found')
            print(err)
        try:
            sheet_name = df_dict['site']
            df_dict['df_feuille_site'].to_excel(writer, sheet_name=sheet_name, index=False)
            writer.save()
            return True
        except Exception as err:
            print('add data to excel file : error {}'.format(err))
            return False

    def treat_data(self, data, fullpath):
        if fullpath:
            self.fullpath = fullpath
        try:
            main_df = get_dataframe_for_final_file()
            pos = 5
            column_with_prices = main_df.columns[pos]
            main_df[column_with_prices] = main_df[column_with_prices].apply(self.clean_euro_random_price_column)
            writer = pd.ExcelWriter(self.fullpath, engine='openpyxl')
            print(f'self.fullpath {self.fullpath}')
            df_feuille1 = main_df.copy(deep=True)
            for spider_name in data:
                print('spider name = {}'.format(spider_name))
                if 'yes' in spider_name or 'rebuy' in spider_name:
                    continue
                df_dict = self.transform_data(spider_name, data[spider_name])
                df_feuille1 = self.concat_data(df_dict, df_feuille1)
                print('excel file created - sheet added {}'.format(spider_name))
            df_feuille1[column_with_prices] = df_feuille1[column_with_prices].apply(lambda x : np.NaN if x == '' or x is None else float(x))
            for spider_name in data:
                df_feuille1[f'{spider_name.upper()}_PrixTTC'] = df_feuille1[f'{spider_name.upper()}_PrixTTC'].apply(lambda x: np.NaN if x == '' or x is None else float(x))
                df_feuille1[f'delta FNAC-DARTY - {spider_name}'] = df_feuille1[column_with_prices] - df_feuille1[f'{spider_name.upper()}_PrixTTC']
            self.add_Feuille1_to_excel_file(df_feuille1, writer)
            writer.close()
            return True
        except Exception as err:
            print('treat data : {}'.format(err))
            writer.close()
            return False

    def concat_data(self, df_dict, df_feuille1):
        df2 = df_feuille1.merge(df_dict['df_feuille1'], on='EAN', how='outer')
        return df2

    def add_Feuille1_to_excel_file(self, df_feuille1, writer):
        try:
            writer.book = load_workbook(self.fullpath)
            writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        except Exception as err:
            print('no file found')
            print(err)
        df_feuille1.drop_duplicates(subset=['EAN'], inplace=True, keep='last')
        final_df = df_feuille1
        final_df.to_excel(writer, index=False, sheet_name='Seconde Vie')
        writer.save()

    @staticmethod
    def match_data_on_ean_one_spreadsheet(df):
        df.columns = df.columns.str.lower()


def main():
    env = 'prod'
    print(env)
    project_script = BaseScrapinghubScript(project_name=spider_names)
    project_script.logger()
    project_script.connect_to_scrapinghub()
    paris_timezone = pytz.timezone('Europe/Paris')
    project_script.run_scrapinghub_spider_list(project_id, spider_names, minutes=2)
    job_keys = project_script.job_keys
    # job_keys = ['436828/23/56', '436828/24/36', '436828/25/28', '436828/30/27', '436828/22/36', '436828/27/29']
    print(job_keys)
    data = project_script.get_scrapinghub_jobs_data(job_keys, spider_names)
    today_ymd = datetime.datetime.now(tz=paris_timezone).strftime("%Y-%m-%d")
    filename = f'{final_filename}_{today_ymd}.xlsx'
    attached_fullpath = f'/tmp/{filename}'
    occ_obj = OccasionsData()
    occ_obj.treat_data(data, attached_fullpath)
    today_dmy = datetime.datetime.now(tz=paris_timezone).strftime("%d/%m/%Y")
    if env == 'prod':
        send_data(mail_recipients,
                  cc_mail_recipients,
                  subject,
                  attached_fullpath,
                  filename,
                  message_text(today_dmy),
                  smtp_access)
    else:
        send_data(['sophie.nassour-ext@fnacdarty.com'],
                  ['sophie.nassour-ext@fnacdarty.com'],
                  subject,
                  attached_fullpath,
                  filename,
                  message_text('DEV !!!'),
                  smtp_access)


if __name__ == '__main__':
    main()
